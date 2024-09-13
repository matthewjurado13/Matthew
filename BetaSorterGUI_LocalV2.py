import sys
from PyQt5.QtWidgets import * 
from PyQt5 import QtGui
import openpyxl
import pandas as pd
import os
from datetime import datetime

class OutputLogger:
    def __init__(self, text_edit):
        self.text_edit = text_edit

    def write(self, message):
        cursor = self.text_edit.textCursor()
        cursor.movePosition(QtGui.QTextCursor.End)
        cursor.insertText(message)
        self.text_edit.setTextCursor(cursor)
        self.text_edit.ensureCursorVisible()

    def flush(self):
        pass

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('GUI Example')
        self.setGeometry(200, 200, 600, 400)
        
        central_widget = QWidget()
        
        # Create a menu bar
        menubar = self.menuBar()
        
        # Create a File menu
        file_menu = menubar.addMenu("File")
        
        # Create a submenu for File menu
        submenu = QMenu("Submenu", self)
        
        # Create actions for the submenu
        action1 = QAction("Action 1", self)
        action2 = QAction("Action 2", self)
        
        # Add actions to the submenu
        submenu.addAction(action1)
        submenu.addAction(action2)

        # Add the submenu to the File menu
        file_menu.addMenu(submenu)
        
        # Add a separator
        file_menu.addSeparator()
        
        # Create an Exit action
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)  
        
        # Create a vertical layout
        layout = QVBoxLayout()
        
        # Create the widgets and add them to the layout
        self.label3 = QLabel('Enter SN to replace:')
        layout.addWidget(self.label3)        

        Sn2replace = QLineEdit()
        Sn2replace.setMaxLength(10)
        Sn2replace.setPlaceholderText("Enter your text")
        Sn2replace.returnPressed.connect(self.return_pressed)
        Sn2replace.selectionChanged.connect(self.selection_changed)
        Sn2replace.textChanged.connect(self.text_changed)
        Sn2replace.textEdited.connect(self.text_edited)
        layout.addWidget(Sn2replace)
        
        self.label4 = QLabel('Stock notes:')
        layout.addWidget(self.label4)        

        notes = QLineEdit()
        notes.setMaxLength(30)
        notes.setPlaceholderText("Enter your text")
        notes.returnPressed.connect(self.return_pressed)
        notes.selectionChanged.connect(self.selection_changed)
        notes.textChanged.connect(self.text_changed)
        notes.textEdited.connect(self.text_edited)
        layout.addWidget(notes)
        
        self.label2 = QLabel('Enter SN:')
        layout.addWidget(self.label2)        

        widget = QLineEdit()
        widget.setMaxLength(10)
        widget.setPlaceholderText("Enter your text")
        widget.returnPressed.connect(self.return_pressed)
        widget.selectionChanged.connect(self.selection_changed)
        widget.textChanged.connect(self.text_changed)
        widget.textEdited.connect(self.text_edited)
        layout.addWidget(widget)
        
        self.label = QLabel('In/Out')
        layout.addWidget(self.label)
        
        self.Incoming = QRadioButton('Incoming')
        self.Incoming.toggled.connect(self.onRadioButtonClicked)
        layout.addWidget(self.Incoming)
        
        self.Outgoing = QRadioButton('Outgoing')
        self.Outgoing.toggled.connect(self.onRadioButtonClicked)
        layout.addWidget(self.Outgoing)
        
        self.combo_box = QComboBox(self)
        self.combo_box.addItem("Beta 1")
        self.combo_box.addItem("Beta 2")
        self.combo_box.addItem("Beta 3")
        self.combo_box.addItem("Beta 4")
        self.combo_box.addItem("Beta 5")
        self.combo_box.addItem("Beta 6")
        self.combo_box.addItem("Beta 7")
        self.combo_box.addItem("Beta 8")
        self.combo_box.addItem("Beta 9")
        self.combo_box.addItem("Beta 10")
        self.combo_box.addItem("Stock")
        self.combo_box.currentIndexChanged.connect(self.on_combobox_changed)
        layout.addWidget(self.combo_box)
        
        self.sortby_combo_box = QComboBox(self)
        self.sortby_combo_box.addItem("Name")
        self.sortby_combo_box.addItem("Model")
        self.sortby_combo_box.addItem("Connector")
        self.sortby_combo_box.addItem("Location")
        self.sortby_combo_box.addItem("Temperature(C)")
        self.sortby_combo_box.addItem("External Calibration Date")
        self.sortby_combo_box.addItem("External Calibration Temperature (C)")
        self.sortby_combo_box.addItem("External Calibration Due Date")
        self.sortby_combo_box.addItem("Self Calibration Date")
        self.sortby_combo_box.addItem("Self Calibration Temperature (C)")
        self.sortby_combo_box.currentIndexChanged.connect(self.on_sortby_combobox_changed)
        layout.addWidget(self.sortby_combo_box)
        
        searchbutton = QPushButton("Search")
        searchbutton.clicked.connect(lambda: self.search_excel(widget.text()))
        layout.addWidget(searchbutton)
        
        button = QPushButton("Enter")
        button.clicked.connect(lambda: self.enter_clicked(widget.text(), Sn2replace.text(), notes.text()))
        layout.addWidget(button)
        
        Sort = QPushButton("Sort")
        Sort.clicked.connect(self.datesortby)
        layout.addWidget(Sort)
        
        Updatebutton = QPushButton("Update")
        Updatebutton.clicked.connect(self.update_excel)
        layout.addWidget(Updatebutton)
        
        calendar = QCalendarWidget(self)
        layout.addWidget(calendar)
        
        # Create the output box
        self.output_box = QTextEdit()
        self.output_box.setReadOnly(True)
        layout.addWidget(self.output_box)

        # Create Clear Button
        clear_button = QPushButton("Clear")
        clear_button.clicked.connect(self.clear_output_box)
        layout.addWidget(clear_button)

        
        
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)
        
        # Redirect print statements to output box
        sys.stdout = OutputLogger(self.output_box)
        sys.stderr = OutputLogger(self.output_box)
        
        self.show()
        cellformat = QtGui.QTextCharFormat()
        cellformat.setBackground(QtGui.QColor("red"))
        df = pd.read_csv(r'allBetas.txt', delimiter="\t")
        dateCol = 'External Calibration Due Date'
        pattern = r'[*]'
        df[dateCol] = df[dateCol].replace(pattern, '', regex=True)
        df[dateCol] = pd.to_datetime(df[dateCol], format='mixed', errors='coerce')
        for i in range(len(df[dateCol])):
            calendar.setDateTextFormat(df[dateCol][i], cellformat)

    def clear_output_box(self):
        self.output_box.clear()

    def onRadioButtonClicked(self):
        global incoming
        selected_option = ''
        if self.Incoming.isChecked():
            selected_option = 'Incoming'
            incoming = 1
        elif self.Outgoing.isChecked():
            selected_option = 'Outgoing'
            incoming = 0
        self.label.setText(f'In/Out: {selected_option}')
        
    def on_combobox_changed(self, index):
        global Beta
        selected_option = self.combo_box.currentText()
        print(f"Selected option: {selected_option}")
        if selected_option == "Stock":
            print(selected_option)
        elif self.combo_box.currentIndex() < 10:
            Beta = self.combo_box.currentIndex() + 1
            print(Beta)
            
    def on_sortby_combobox_changed(self, index):
        selected_option = self.sortby_combo_box.currentText()
        print(f"Selected option: {selected_option}")
        if selected_option == "Name":
            print(selected_option)
        elif selected_option == "Model":
            print(selected_option)
        elif selected_option == "Connector":
            print(selected_option)
        elif selected_option == "Location":
            print(selected_option)
        elif selected_option == "Temperature(C)":
            print(selected_option)            
        elif selected_option == "External Calibration Date":
            print(selected_option)            
        elif selected_option == "External Calibration Temperature (C)":
            print(selected_option)            
        elif selected_option == "External Calibration Due Date":
            print(selected_option)
        elif selected_option == "Self Calibration Date":
            print(selected_option)
        elif selected_option == "Self Calibration Temperature (C)":
            print(selected_option)

    def return_pressed(self):
        pass        
    def selection_changed(self):
        pass
    def text_changed(self, s):
        pass
    def text_edited(self, s):
       pass
    
    def enter_clicked(self, newSN, oldSN, notes):
        print("Enter was clicked")
        try:
            workbook = openpyxl.load_workbook(r'BetaSysCardSummariesTest.xlsx')
            selected_option = self.combo_box.currentText()
            
            if selected_option == "Stock":
                worksheet = workbook[selected_option]
                print(selected_option)
                next_row = 1
                while worksheet.cell(row=next_row, column=1).value is not None:
                    next_row += 1
                
                worksheet.cell(row=next_row, column=1, value=newSN)
                worksheet.cell(row=next_row, column=2, value=oldSN)
                worksheet.cell(row=next_row, column=3, value=notes)
                
                workbook.save(r'BetaSysCardSummariesTest.xlsx')
                print(f"Added new row to {selected_option} worksheet")
            else:
                worksheet = workbook[f"Beta {self.combo_box.currentIndex() + 1}"]
                print(f"Selected Beta {self.combo_box.currentIndex() + 1}")
                next_row = 1
                while worksheet.cell(row=next_row, column=1).value is not None:
                    next_row += 1
                
                worksheet.cell(row=next_row, column=1, value=newSN)
                worksheet.cell(row=next_row, column=2, value=oldSN)
                worksheet.cell(row=next_row, column=3, value=notes)
                
                workbook.save(r'BetaSysCardSummariesTest.xlsx')
                print(f"Added new row to Beta {self.combo_box.currentIndex() + 1} worksheet")

        except Exception as e:
            print(f"An error occurred: {e}")

    def search_excel(self,search_value):
        try:
            workbook = openpyxl.load_workbook(r'BetaSysCardSummariesTest.xlsx')
            found=0        
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value == search_value:
                            print(f"Found at cell {cell.coordinate}: {cell.value}")
                            print("Worksheet is:"+str(worksheet))
                            found+=1                        
            if found == 0:
                print("nothing found")
        except:
            print("Couldn't open Excel file")
        
    def update_excel(self):
        print("Updating Excel File")
        try:
            df = pd.read_excel(r'BetaSysCardSummariesTest.xlsx', sheet_name=None)
            with pd.ExcelWriter(r'BetaSysCardSummariesTest.xlsx', engine='openpyxl') as writer:
                for sheet_name, sheet_df in df.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print("Excel file updated successfully.")
        except Exception as e:
            print(f"An error occurred: {e}")

    def datesortby(self):
        print("Sorting by date...")
        try:
            # Load the data
            df = pd.read_csv(r'allBetas.txt', delimiter="\t")
            date_col = 'External Calibration Due Date'
            df[date_col] = pd.to_datetime(df[date_col], format='%Y-%m-%d', errors='coerce')
            sorted_df = df.sort_values(by=date_col)

            # Save the sorted data
            sorted_df.to_csv(r'allBetas_sorted.txt', index=False, sep='\t')
            print("Data sorted and saved to 'allBetas_sorted.txt'")
        except Exception as e:
            print(f"An error occurred: {e}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainWindow()
    sys.exit(app.exec_())


